import os, re
import pandas as pd
from typing import Tuple, List

VALID_EXT = {".xlsx", ".xls", ".csv"}

def ensure_dirs():
    os.makedirs("uploads", exist_ok=True)
    os.makedirs("exports", exist_ok=True)

def safe_table_name(name: str) -> str:
    import os, re
    base = os.path.splitext(os.path.basename(name))[0]
    base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_").lower()
    if not base:
        base = "imported_table"
    if not re.match(r"^[a-zA-Z]", base):
        base = "tbl_" + base
    return base

def read_any(path: str, sheet_name=None) -> pd.DataFrame | dict:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    else:
        # if no sheet specified, we still allow returning a specific sheet (handled by caller)
        if sheet_name is None:
            sheet_name = 0  # default to first sheet
        return pd.read_excel(path, sheet_name=sheet_name)

def list_sheets(path: str):
    """Return sheet names for Excel files, or None for CSVs."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in {".xlsx", ".xls"}:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        return wb.sheetnames
    except Exception as e:
        print("Error reading sheet names:", e)
        return None

def sanitize_headers(cols: List[str]) -> List[str]:
    clean = []
    for i, c in enumerate(cols):
        c = str(c).strip()
        if not c or c.lower() == "nan":
            c = f"col_{i+1}"
        c = re.sub(r"\s+", "_", c)
        c = re.sub(r"[^A-Za-z0-9_]", "", c)
        if re.match(r"^\d", c):
            c = f"c_{c}"
        clean.append(c.lower())
    seen = {}
    final = []
    for c in clean:
        if c not in seen:
            seen[c] = 1
            final.append(c)
        else:
            seen[c] += 1
            final.append(f"{c}_{seen[c]}")
    return final

def pandas_to_sqlite_types(df: pd.DataFrame) -> Tuple[str, list]:
    col_defs = []
    cols = list(df.columns)
    for c in cols:
        dt = str(df[c].dtype)
        if "int" in dt:
            sqlt = "INTEGER"
        elif "float" in dt:
            sqlt = "REAL"
        elif "bool" in dt:
            sqlt = "INTEGER"
        elif "datetime" in dt:
            sqlt = "TEXT"
            df[c] = pd.to_datetime(df[c], errors="coerce").astype("datetime64[ns]")
            df[c] = df[c].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
        else:
            sqlt = "TEXT"
            df[c] = df[c].astype(str)
        col_defs.append(f"{c} {sqlt}")
    if "id" not in cols:
        col_defs.insert(0, "id INTEGER PRIMARY KEY AUTOINCREMENT")
    return ", ".join(col_defs), cols

def df_rows(df: pd.DataFrame, cols: list):
    clean = df.where(pd.notnull(df), None)
    return clean[cols].itertuples(index=False, name=None)

def df_html(df: pd.DataFrame, max_rows=100) -> str:
    return df.head(max_rows).to_html(classes="table", index=False, border=0, escape=False)
